"""
Export results module — HTML report, Excel, PNG charts.
"""

import io
import base64
import os
from datetime import datetime
from i18n import tr, get_language


def _fig_to_base64(fig) -> str:
    """Convert a matplotlib figure to base64 PNG string."""
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=200, facecolor='#1e1e1e', edgecolor='none',
                bbox_inches='tight')
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


def _format_pvalue(p: float) -> str:
    if p < 0.0001:
        return "p < 0.0001"
    elif p < 0.001:
        return f"p = {p:.4f}"
    elif p < 0.01:
        return f"p = {p:.5f}"
    else:
        return f"p = {p:.5f}"


def _strength_label(s: str) -> str:
    return s.replace('_', ' ').capitalize()


def export_html(results: dict, d1, d2, name1: str, name2: str,
                output_path: str) -> str:
    """Generate a self-contained HTML report with all tables and charts."""

    lang = get_language()
    desc1 = results["descriptive1"]
    desc2 = results["descriptive2"]
    pearson = results["pearson"]
    spearman = results["spearman"]
    reg = results["regression_2on1"]
    ind = results["independent_ttest"]
    paired = results["paired_ttest"]
    os1 = results["onesample_ttest1"]
    os2 = results["onesample_ttest2"]
    n1 = results["normality1"]
    n2 = results["normality2"]
    dc = results["distribution_comparison"]

    # Generate chart images
    from charts import (
        create_histogram_chart, create_boxplot_chart, create_scatter_chart,
        create_qq_side_by_side, create_residuals_chart
    )
    n = min(len(d1), len(d2))
    hist = create_histogram_chart(d1, d2, name1, name2)
    box = create_boxplot_chart(d1, d2, name1, name2)
    scatter = create_scatter_chart(d1[:n], d2[:n], name1, name2,
                                   slope=reg['slope'], intercept=reg['intercept'],
                                   r=pearson['r'])
    qq = create_qq_side_by_side(d1, d2, name1, name2)
    resid = create_residuals_chart(reg['residuals'], reg['predictions'])

    img_hist = _fig_to_base64(hist.fig)
    img_box = _fig_to_base64(box.fig)
    img_scatter = _fig_to_base64(scatter.fig)
    img_qq = _fig_to_base64(qq.fig)
    img_resid = _fig_to_base64(resid.fig)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="{lang}">
<head>
<meta charset="UTF-8">
<title>Statistical Analysis Report</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #1e1e1e; color: #d4d4d4;
         max-width: 1100px; margin: 0 auto; padding: 30px; }}
  h1 {{ color: #569cd6; text-align: center; margin-bottom: 5px; }}
  h2 {{ color: #569cd6; border-bottom: 1px solid #3c3c3c; padding-bottom: 6px; margin-top: 30px; }}
  h3 {{ color: #dcdcaa; margin-top: 25px; }}
  .subtitle {{ text-align: center; color: #888; margin-bottom: 30px; }}
  table {{ width: 100%; border-collapse: collapse; margin: 10px 0 20px 0; }}
  th {{ background: #2d2d2d; color: #d4d4d4; padding: 8px 12px; text-align: left;
        border: 1px solid #3c3c3c; font-weight: bold; }}
  td {{ padding: 6px 12px; border: 1px solid #3c3c3c; }}
  tr:nth-child(even) {{ background: #252526; }}
  .sig {{ color: #4ec9b0; font-weight: bold; }}
  .notsig {{ color: #ce9178; font-weight: bold; }}
  .comment {{ background: #252526; border-left: 3px solid #569cd6; padding: 10px 15px;
             margin: 10px 0; color: #aaa; font-style: italic; }}
  img {{ width: 100%; margin: 10px 0; border: 1px solid #3c3c3c; border-radius: 4px; }}
  .chart-container {{ aspect-ratio: 16 / 9; width: 100%; overflow: hidden;
                      background: #252526; border: 1px solid #3c3c3c; border-radius: 4px;
                      margin: 15px 0; display: flex; align-items: center; justify-content: center; }}
  .chart-container img {{ width: 100%; height: 100%; object-fit: contain; margin: 0; border: none; }}
  .footer {{ text-align: center; color: #666; margin-top: 50px; font-size: 14px; }}
  .credit {{ color: #aaa; font-weight: bold; }}
</style>
</head>
<body>
<h1>{tr("app_title")}</h1>
<p class="subtitle">Generated: {now} &nbsp;|&nbsp; {name1} ({desc1['n']} values) vs {name2} ({desc2['n']} values)</p>

<h2>{tr("desc_title")}</h2>
<table>
<tr><th>{tr("desc_col_stat")}</th><th>{name1}</th><th>{name2}</th></tr>
<tr><td>{tr("row_n")}</td><td>{desc1['n']}</td><td>{desc2['n']}</td></tr>
<tr><td>{tr("row_mean")}</td><td>{desc1['mean']:.6f}</td><td>{desc2['mean']:.6f}</td></tr>
<tr><td>{tr("row_median")}</td><td>{desc1['median']:.6f}</td><td>{desc2['median']:.6f}</td></tr>
<tr><td>{tr("row_std")}</td><td>{desc1['std']:.6f}</td><td>{desc2['std']:.6f}</td></tr>
<tr><td>{tr("row_variance")}</td><td>{desc1['variance']:.6f}</td><td>{desc2['variance']:.6f}</td></tr>
<tr><td>{tr("row_skewness")}</td><td>{desc1['skewness']:.4f}</td><td>{desc2['skewness']:.4f}</td></tr>
<tr><td>{tr("row_kurtosis")}</td><td>{desc1['kurtosis']:.4f}</td><td>{desc2['kurtosis']:.4f}</td></tr>
<tr><td>{tr("row_range")}</td><td>{desc1['range']:.6f}</td><td>{desc2['range']:.6f}</td></tr>
<tr><td>{tr("row_iqr")}</td><td>{desc1['iqr']:.6f}</td><td>{desc2['iqr']:.6f}</td></tr>
<tr><td>{tr("row_ci_mean")}</td><td>[{desc1['ci_95_lower']:.4f}, {desc1['ci_95_upper']:.4f}]</td><td>[{desc2['ci_95_lower']:.4f}, {desc2['ci_95_upper']:.4f}]</td></tr>
</table>

<h2>{tr("tab_correlations")}</h2>
<h3>{tr("corr_pearson_title")}</h3>
<table>
<tr><th>{tr("corr_col_measure")}</th><th>{tr("corr_col_value")}</th></tr>
<tr><td>r</td><td>{pearson['r']:.6f}</td></tr>
<tr><td>r²</td><td>{pearson['r2']:.6f}</td></tr>
<tr><td>p</td><td>{_format_pvalue(pearson['p_value'])}</td></tr>
<tr><td>95% CI</td><td>[{pearson['ci_95_lower']:.4f}, {pearson['ci_95_upper']:.4f}]</td></tr>
<tr><td>{tr("corr_strength")}</td><td>{_strength_label(pearson['strength'])}</td></tr>
<tr><td>{tr("corr_conclusion")}</td><td class="{'sig' if pearson['significant'] else 'notsig'}">{tr('corr_significant') if pearson['significant'] else tr('corr_not_significant')}</td></tr>
</table>

<h3>{tr("corr_spearman_title")}</h3>
<table>
<tr><th>{tr("corr_col_measure")}</th><th>{tr("corr_col_value")}</th></tr>
<tr><td>ρ</td><td>{spearman['rho']:.6f}</td></tr>
<tr><td>p</td><td>{_format_pvalue(spearman['p_value'])}</td></tr>
<tr><td>{tr("corr_strength")}</td><td>{_strength_label(spearman['strength'])}</td></tr>
</table>

<h2>{tr("tab_regression")}</h2>
<p style="font-family:monospace;color:#dcdcaa;">{reg['equation']}</p>
<table>
<tr><th>{tr("desc_col_stat")}</th><th>{tr("corr_col_value")}</th></tr>
<tr><td>{tr("reg_r2")}</td><td>{reg['r2']:.6f}</td></tr>
<tr><td>{tr("reg_r2_adj")}</td><td>{reg['r2_adj']:.6f}</td></tr>
<tr><td>{tr("reg_f")}</td><td>F(1, {reg['df_residual']}) = {reg['f_statistic']:.4f}</td></tr>
<tr><td>{tr("reg_f_pval")}</td><td>{_format_pvalue(reg['f_pvalue'])}</td></tr>
<tr><td>{tr("reg_intercept")}</td><td>{reg['intercept']:.6f} (SE={reg['intercept_se']:.6f}, p={_format_pvalue(reg['intercept_p'])})</td></tr>
<tr><td>{tr("reg_slope")}</td><td>{reg['slope']:.6f} (SE={reg['slope_se']:.6f}, p={_format_pvalue(reg['slope_p'])})</td></tr>
</table>

<h2>{tr("tab_ttests")}</h2>
<h3>{tr("ttest_ind_title")}</h3>
<table>
<tr><th>{tr("ttest_col_stat")}</th><th>{tr("ttest_col_student")}</th><th>{tr("ttest_col_welch")}</th></tr>
<tr><td>{tr("ttest_t")}</td><td>{ind['student_t']:.4f}</td><td>{ind['welch_t']:.4f}</td></tr>
<tr><td>{tr("ttest_df")}</td><td>{ind['student_df']}</td><td>≈ {round((ind['welch_t']**2)/(ind['mean_difference']**2)*(ind['n1']+ind['n2']-2) if ind['welch_t']!=0 else 0, 1)}</td></tr>
<tr><td>{tr("ttest_pvalue")}</td><td>{_format_pvalue(ind['student_pvalue'])}</td><td>{_format_pvalue(ind['welch_pvalue'])}</td></tr>
</table>
<p>Mean difference: {ind['mean_difference']:.4f} (95% CI: [{ind['ci_95_lower']:.4f}, {ind['ci_95_upper']:.4f}]) | Cohen's d = {ind['cohens_d']:.4f} ({_strength_label(ind['cohens_d_strength'])})</p>

<h3>{tr("ttest_paired_title")}</h3>
<table>
<tr><td>N (pairs)</td><td>{paired['n_pairs']}</td></tr>
<tr><td>Mean difference</td><td>{paired['mean_difference']:.4f}</td></tr>
<tr><td>t</td><td>t({paired['df']}) = {paired['t_statistic']:.4f}</td></tr>
<tr><td>p</td><td>{_format_pvalue(paired['p_value'])}</td></tr>
<tr><td>Cohen's d</td><td>{paired['cohens_d']:.4f}</td></tr>
</table>

<h2>{tr("tab_normality")}</h2>
<table>
<tr><th>Test</th><th>{name1}</th><th>{name2}</th></tr>
<tr><td>Shapiro-Wilk p</td><td>{_format_pvalue(n1['shapiro_pvalue']) if n1['shapiro_pvalue'] else 'N/A'}</td><td>{_format_pvalue(n2['shapiro_pvalue']) if n2['shapiro_pvalue'] else 'N/A'}</td></tr>
<tr><td>D'Agostino p</td><td>{_format_pvalue(n1['dagostino_pvalue'])}</td><td>{_format_pvalue(n2['dagostino_pvalue'])}</td></tr>
</table>

<h2>{tr("tab_distributions")}</h2>
<table>
<tr><td>KS D</td><td>{dc['ks_statistic']:.4f}</td></tr>
<tr><td>KS p</td><td>{_format_pvalue(dc['ks_pvalue'])}</td></tr>
<tr><td>Mann-Whitney U</td><td>{dc['mannwhitney_u']:.4f}</td></tr>
<tr><td>MW p</td><td>{_format_pvalue(dc['mannwhitney_pvalue'])}</td></tr>
</table>

<h2>{tr("tab_charts")}</h2>
<div class="chart-container"><img src="data:image/png;base64,{img_hist}" alt="Histograms"></div>
<div class="chart-container"><img src="data:image/png;base64,{img_box}" alt="Box Plots"></div>
<div class="chart-container"><img src="data:image/png;base64,{img_scatter}" alt="Scatter Plot"></div>
<div class="chart-container"><img src="data:image/png;base64,{img_qq}" alt="Q-Q Plots"></div>
<div class="chart-container"><img src="data:image/png;base64,{img_resid}" alt="Residuals"></div>

<div class="footer">
<p class="credit">{tr("app_credit")}</p>
<p>Generated by Statistics Tool</p>
</div>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return output_path


def export_excel(results: dict, name1: str, name2: str, output_path: str) -> str:
    """Export all result tables to an Excel file with multiple sheets."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Descriptive"

    desc1 = results["descriptive1"]
    desc2 = results["descriptive2"]
    pearson = results["pearson"]
    spearman = results["spearman"]
    reg = results["regression_2on1"]
    ind = results["independent_ttest"]
    paired = results["paired_ttest"]
    n1 = results["normality1"]
    n2 = results["normality2"]
    dc = results["distribution_comparison"]

    # Descriptive
    headers = ["Statistic", name1, name2]
    ws.append(headers)
    for label, key1, key2 in [
        ("N", "n", "n"), ("Mean", "mean", "mean"), ("Median", "median", "median"),
        ("Std Dev", "std", "std"), ("Variance", "variance", "variance"),
        ("Skewness", "skewness", "skewness"), ("Kurtosis", "kurtosis", "kurtosis"),
        ("Min", "min", "min"), ("Max", "max", "max"), ("Range", "range", "range"),
        ("Q1", "q1", "q1"), ("Q3", "q3", "q3"), ("IQR", "iqr", "iqr"),
        ("95% CI Lower", "ci_95_lower", "ci_95_lower"),
        ("95% CI Upper", "ci_95_upper", "ci_95_upper"),
    ]:
        ws.append([label, desc1[key1], desc2[key2]])

    # Correlations
    ws = wb.create_sheet("Correlations")
    ws.append(["Measure", "Value"])
    ws.append(["Pearson r", pearson['r']])
    ws.append(["Pearson r²", pearson['r2']])
    ws.append(["Pearson p", pearson['p_value']])
    ws.append(["Pearson 95% CI Lower", pearson['ci_95_lower']])
    ws.append(["Pearson 95% CI Upper", pearson['ci_95_upper']])
    ws.append(["Spearman ρ", spearman['rho']])
    ws.append(["Spearman p", spearman['p_value']])

    # Regression
    ws = wb.create_sheet("Regression")
    ws.append(["Statistic", "Value"])
    ws.append(["Equation", reg['equation']])
    ws.append(["R²", reg['r2']])
    ws.append(["Adjusted R²", reg['r2_adj']])
    ws.append(["F-statistic", reg['f_statistic']])
    ws.append(["F p-value", reg['f_pvalue']])
    ws.append(["Intercept", reg['intercept']])
    ws.append(["Intercept SE", reg['intercept_se']])
    ws.append(["Intercept p", reg['intercept_p']])
    ws.append(["Slope", reg['slope']])
    ws.append(["Slope SE", reg['slope_se']])
    ws.append(["Slope p", reg['slope_p']])
    ws.append(["MSE", reg['mse']])

    # t-Tests
    ws = wb.create_sheet("t-Tests")
    ws.append(["Test", "Statistic", "Value"])
    ws.append(["Student's t", "t", ind['student_t']])
    ws.append(["Student's t", "df", ind['student_df']])
    ws.append(["Student's t", "p", ind['student_pvalue']])
    ws.append(["Welch's t", "t", ind['welch_t']])
    ws.append(["Welch's t", "p", ind['welch_pvalue']])
    ws.append(["Levene", "F", ind['levene_statistic']])
    ws.append(["Levene", "p", ind['levene_pvalue']])
    ws.append(["Cohen's d", "d", ind['cohens_d']])
    ws.append(["Paired t", "t", paired['t_statistic']])
    ws.append(["Paired t", "df", paired['df']])
    ws.append(["Paired t", "p", paired['p_value']])
    ws.append(["Paired Cohen's d", "d", paired['cohens_d']])

    # Normality
    ws = wb.create_sheet("Normality")
    ws.append(["Test", name1, name2])
    ws.append(["Shapiro-Wilk p", n1['shapiro_pvalue'] or "N/A", n2['shapiro_pvalue'] or "N/A"])
    ws.append(["D'Agostino p", n1['dagostino_pvalue'], n2['dagostino_pvalue']])

    # Distribution
    ws = wb.create_sheet("Distribution")
    ws.append(["Test", "Value"])
    ws.append(["KS D", dc['ks_statistic']])
    ws.append(["KS p", dc['ks_pvalue']])
    ws.append(["Mann-Whitney U", dc['mannwhitney_u']])
    ws.append(["MW p", dc['mannwhitney_pvalue']])

    wb.save(output_path)
    return output_path


def export_charts_png(results: dict, d1, d2, name1: str, name2: str,
                      output_dir: str) -> list:
    """Export all charts as separate PNG files. Returns list of paths."""
    from charts import (
        create_histogram_chart, create_boxplot_chart, create_scatter_chart,
        create_qq_side_by_side, create_residuals_chart
    )

    os.makedirs(output_dir, exist_ok=True)
    reg = results["regression_2on1"]
    pearson = results["pearson"]
    n = min(len(d1), len(d2))

    charts = [
        ("histogram", create_histogram_chart(d1, d2, name1, name2)),
        ("boxplot", create_boxplot_chart(d1, d2, name1, name2)),
        ("scatter", create_scatter_chart(d1[:n], d2[:n], name1, name2,
                                         slope=reg['slope'], intercept=reg['intercept'],
                                         r=pearson['r'])),
        ("qq_plot", create_qq_side_by_side(d1, d2, name1, name2)),
        ("residuals", create_residuals_chart(reg['residuals'], reg['predictions'])),
    ]

    paths = []
    for name, chart in charts:
        path = os.path.join(output_dir, f"{name}.png")
        chart.fig.savefig(path, dpi=150, facecolor='#1e1e1e', edgecolor='none',
                          bbox_inches='tight')
        paths.append(path)

    return paths


def export_all(results: dict, d1, d2, name1: str, name2: str,
               output_dir: str) -> dict:
    """Export everything: HTML report, Excel, and PNG charts.
    Returns dict with paths."""
    os.makedirs(output_dir, exist_ok=True)

    paths = {}

    paths['html'] = export_html(results, d1, d2, name1, name2,
                                os.path.join(output_dir, "report.html"))

    paths['excel'] = export_excel(results, name1, name2,
                                  os.path.join(output_dir, "results.xlsx"))

    paths['charts'] = export_charts_png(results, d1, d2, name1, name2,
                                        os.path.join(output_dir, "charts"))

    return paths